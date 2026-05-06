#!/usr/bin/env python3
"""Generates Pumplo Instagram Marketing Strategy Excel for the whole team."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import date, timedelta
import os

# ── Brand colors ──────────────────────────────────────────────────────────────
NAVY       = "0B1222"
BLUE       = "4CC9FF"
WHITE      = "FFFFFF"
GRAY_LIGHT = "F4F6F8"
GRAY_MID   = "D0D7E2"
GRAY_DARK  = "6B7A90"
GREEN      = "34C759"
ORANGE     = "FF9F0A"
RED        = "FF3B30"
PURPLE     = "AF52DE"
YELLOW     = "FFD60A"

# content type colors
COLOR_EDU    = "C8F5E0"  # vzdělávání - green tint
COLOR_HUM    = "FFF3C8"  # humor - yellow tint
COLOR_APP    = "C8E8FF"  # appka/social proof - blue tint
COLOR_BTS    = "F5C8FF"  # behind the scenes - purple tint
COLOR_OFF    = "F0F0F0"  # day off / stories only

# phase colors
PHASE_COLORS = {
    "Fáze 0 — Rozjezd":       "FFE4B5",
    "Fáze 0b — Teaser/FOMO":  "FFCC80",
    "Fáze 1 — Launch blast":   "FFB347",
    "Fáze 1b — Post-launch":   "FFF176",
    "Fáze 2 — Momentum":       "90EE90",
    "Fáze 3 — Škálování":      "87CEEB",
    "Fáze 4 — Sprint":         "DDA0DD",
}

COLOR_LAUNCH = "FF8A65"   # warm orange — launch posts
COLOR_TEASER = "FFCC80"   # amber — teaser/FOMO posts

def fill(hex_color, fill_type="solid"):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type=fill_type)

def bold_font(color=None, size=11):
    return Font(bold=True, color=color or "000000", size=size)

def header_font(size=12):
    return Font(bold=True, color=WHITE, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color=GRAY_MID)
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def header_row(ws, row, headers, bg=NAVY, font_color=WHITE, height=30):
    set_row_height(ws, row, height)
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill(bg)
        cell.font = Font(bold=True, color=font_color, size=10)
        cell.alignment = center()
        cell.border = thin_border()

def section_title(ws, row, col, text, colspan=1, bg=NAVY):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color=WHITE, size=13)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + colspan - 1)
    set_row_height(ws, row, 28)

def data_cell(ws, row, col, value, bg=WHITE, font_color="000000",
              bold=False, align="left", wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=font_color, size=10)
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    cell.border = thin_border()
    return cell

# ── Special day overrides: (date) -> (fmt, typ, tema, content_color, poznamka) ─
SPECIAL_DAYS = {
    # ── Úvodní post ──────────────────────────────────────────────────────────
    date(2026, 5, 11): (
        "Post", "Úvodní post", "✅ HOTOVO — post připravený přes marketing bot na Telegramu. Vysvětluje celou appku. Zkopírovat caption z Telegramu a zveřejnit.", COLOR_BTS,
        "⭐ PRVNÍ PŘÍSPĚVEK — caption a vizuál již připraven marketing botem na Telegramu. Jen zkopírovat a zveřejnit přes Ayrshare."
    ),

    # ── Teaser / FOMO týden (25.–31. 5.) ─────────────────────────────────────
    date(2026, 5, 25): (
        "Reel", "Teaser", "👀 'Něco přichází...' — zákulisí přípravy samolepek a gymů (bez odhalení)", COLOR_TEASER,
        "FOMO teaser — neodhaluj víc než nutné"
    ),
    date(2026, 5, 26): (
        "Stories", "Countdown", "⏳ Countdown stories: 7 dní do launche — poll 'Uhádneš co to je?'", COLOR_TEASER,
        "Jen Stories, ne feed post"
    ),
    date(2026, 5, 27): (
        "Carousel", "Teaser", "🏋️ 'Tyhle posilovny brzy změní způsob, jak cvičíš' — 5 gymů bez jmen", COLOR_TEASER,
        "Budovat napětí — gymy anonymně"
    ),
    date(2026, 5, 28): (
        "Reel", "Teaser", "🤫 'Dáváme fitku superschopnosti' — záběry z Eurogymu bez odhalení appky", COLOR_TEASER,
        "Virální potenciál — mysteriózní tón"
    ),
    date(2026, 5, 29): (
        "Reel", "Teaser", "📱 '3 dny...' — rychlý střih app screenshotů, hudba, žádný text", COLOR_TEASER,
        "App sneak peek — jen vizuály, žádné vysvětlení"
    ),
    date(2026, 5, 30): (
        "Reel", "Teaser", "🔥 'Zítra.' — David v gymu, kamera za ním, minimalistický text", COLOR_TEASER,
        "Den před launchem — maximální napětí"
    ),
    date(2026, 5, 31): (
        "Stories", "Countdown", "⏰ 'Zítra to jde live!' — countdown sticker + swipe up na profil", COLOR_TEASER,
        "Jen Stories — finální odpočet"
    ),

    # ── Launch týden (1.–7. 6.) ───────────────────────────────────────────────
    date(2026, 6, 1): (
        "Reel", "🚀 LAUNCH", "🎉 LAUNCH DAY: 'Pumplo je live! Najdeš nás v těchto 5 posilovnách' — David mluví do kamery, ukáže samolepky", COLOR_LAUNCH,
        "🚀 LAUNCH DAY — připravit 3 dny předem, David schvaluje osobně, zveřejnit 8:00"
    ),
    date(2026, 6, 2): (
        "Carousel", "🚀 Launch", "'Kde všude teď cvičíš s Pumplo?' — 1 slide = 1 gym, fotka + adresa + jak se zapojit", COLOR_LAUNCH,
        "Tag každého gymu v příspěvku — collab funkce IG"
    ),
    date(2026, 6, 3): (
        "Reel", "🚀 Launch", "'Jak to funguje? 60 sekund a víš vše' — screen recording app + komentář Davida", COLOR_LAUNCH,
        "App walkthrough — jasné, jednoduché, žádný jargon"
    ),
    date(2026, 6, 4): (
        "Carousel", "🚀 Launch", "'Jak začít s Pumplo? 3 jednoduché kroky' — 1) jdi do gymu 2) naskenuj QR 3) trénuj podle plánu", COLOR_LAUNCH,
        "Onboarding guide — sdílitelné, pin na profil"
    ),
    date(2026, 6, 5): (
        "Reel", "🚀 Launch", "'První členové mluví' — krátké záběry z gymů, členové u appky, trenér komentuje", COLOR_LAUNCH,
        "Natočit 1.–4. 6. v gymech — testimonials v akci"
    ),
    date(2026, 6, 6): (
        "Reel", "🚀 Collab", "'Eurogym Olomouc × Pumplo' — collab Reel s Eurogymem, trenér ukazuje appku členům", COLOR_LAUNCH,
        "IG Collab post s @eurogym profilem — dosah na jejich sledující"
    ),
    date(2026, 6, 7): (
        "Carousel", "🚀 Launch recap", "'Týden s Pumplo — co se stalo za 7 dní' — čísla, gymy, reakce členů", COLOR_LAUNCH,
        "Launch recap — pin na profil, sdílej na stories"
    ),
}

# ── Helper: get phase from date ────────────────────────────────────────────────
def get_phase(d: date) -> str:
    if d < date(2026, 5, 25):
        return "Fáze 0 — Rozjezd"
    elif d < date(2026, 6, 1):
        return "Fáze 0b — Teaser/FOMO"
    elif d < date(2026, 6, 8):
        return "Fáze 1 — Launch blast"
    elif d < date(2026, 7, 1):
        return "Fáze 1b — Post-launch"
    elif d < date(2026, 9, 1):
        return "Fáze 2 — Momentum"
    elif d < date(2026, 11, 1):
        return "Fáze 3 — Škálování"
    else:
        return "Fáze 4 — Sprint"

# ── Weekly schedule definition ─────────────────────────────────────────────────
CZ_DAYS = ["Pondělí","Úterý","Středa","Čtvrtek","Pátek","Sobota","Neděle"]

WEEKLY = {
    0: ("Reel",    "Vzdělávání",   "Workout tip nebo fitness fakt",          COLOR_EDU),
    1: ("Carousel","Vzdělávání",   "5 tipů, checklist nebo mýtus vs. pravda",COLOR_EDU),
    2: ("Reel",    "Humor",        "Relatable gymová situace",               COLOR_HUM),
    3: ("Carousel","Appka/Social proof","App feature nebo screenshoty z gymu",COLOR_APP),
    4: ("Reel",    "Vzdělávání/Humor","Volba týmu — co teď letí",           COLOR_EDU),
    5: ("Reel",    "Collab/B2B",   "Gymový collab nebo obsah pro majitele",  COLOR_APP),
    6: (None,      "Stories only", "Zákulisí, poll, Q&A — žádný feed post",  COLOR_OFF),
}

# Fáze 0: only Mon/Wed/Fri Reels (3/week), no Saturday Reel
PHASE0_SKIP = {1, 3, 5}  # Úterý, Čtvrtek, Sobota = no post in phase 0

def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    make_strategie(wb)
    make_kalendar(wb)
    make_backlog(wb)
    make_hashtagy(wb)
    make_kpis(wb)
    make_bot_guide(wb)

    return wb

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — STRATEGIE
# ══════════════════════════════════════════════════════════════════════════════
def make_strategie(wb):
    ws = wb.create_sheet("📊 Strategie")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 18

    r = 1

    # ── Title ──
    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1, value="PUMPLO — INSTAGRAM MARKETINGOVÁ STRATEGIE 2026")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=16)
    c.alignment = center()
    set_row_height(ws, r, 42)
    r += 1

    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1,
        value="Cíl: 10 000 followers do 31. 12. 2026  |  Start: 15 followers, 0 příspěvků  |  Budget: 0 Kč (organicky)")
    c.fill = fill(BLUE)
    c.font = Font(bold=True, color=NAVY, size=11)
    c.alignment = center()
    set_row_height(ws, r, 24)
    r += 2

    # ── CONTENT MIX ──
    section_title(ws, r, 1, "  CONTENT MIX", 4)
    r += 1
    header_row(ws, r, ["Typ obsahu", "Podíl %", "Popis a příklady", "Barva v kalendáři"])
    r += 1

    mix_data = [
        ("Vzdělávání",        "40 %",
         "Workout tipy, fitness fakta, how-to, jak funguje retence, checklist", COLOR_EDU),
        ("Humor / Zábava",    "30 %",
         "Relatable gymové situace, memes, 'ty to znáš' momenty", COLOR_HUM),
        ("Appka + Social proof","20 %",
         "App features, screenshoty z gymu, členové v akci, Eurogym case study", COLOR_APP),
        ("Behind the scenes", "10 %",
         "David v gymech, tým, příprava launche, trenéři, zákulisí", COLOR_BTS),
    ]
    for typ, podil, popis, color in mix_data:
        set_row_height(ws, r, 36)
        data_cell(ws, r, 1, typ, bg=color, bold=True)
        data_cell(ws, r, 2, podil, bg=color, bold=True, align="center")
        data_cell(ws, r, 3, popis, bg=color)
        data_cell(ws, r, 4, "████", bg=color, align="center", font_color=color)
        r += 1
    r += 1

    # ── FÁZE A MILNÍKY ──
    section_title(ws, r, 1, "  FÁZE A MILNÍKY", 4)
    r += 1
    header_row(ws, r, ["Fáze", "Období", "Hlavní páka", "Cíl followers"])
    r += 1

    faze_data = [
        ("Fáze 0 — Rozjezd",     "Květen 2026",       "Prvních 9–12 postů, FOMO před launchem, vyplnit grid",              "200",   "FFE4B5"),
        ("Fáze 1 — Launch blast", "Červen 2026",       "5 gymů + QR samolepky, collab posty s trenéry, launch obsah",       "600",   "FFB347"),
        ("Fáze 2 — Momentum",     "Červenec–Srpen",    "5 Reels/týden, Simona rozjetá na CapCut, první virální Reel",       "1 500", "90EE90"),
        ("Fáze 3 — Škálování",    "Září–Říjen",        "Case study Eurogym, UGC od členů, opakovat co funguje",             "4 000", "87CEEB"),
        ("Fáze 4 — Sprint",       "Listopad–Prosinec", "Vánoční kampaň, membership jako dárek, B2B push, Instagram SEO",    "10 000","DDA0DD"),
    ]
    for faze, obdobi, paka, cil, color in faze_data:
        set_row_height(ws, r, 38)
        data_cell(ws, r, 1, faze, bg=color, bold=True)
        data_cell(ws, r, 2, obdobi, bg=color, align="center")
        data_cell(ws, r, 3, paka, bg=color)
        data_cell(ws, r, 4, cil, bg=color, bold=True, align="center")
        r += 1
    r += 1

    # ── POSTING SCHEDULE ──
    section_title(ws, r, 1, "  TÝDENNÍ POSTING SCHEDULE", 4)
    r += 1
    header_row(ws, r, ["Den", "Formát", "Typ obsahu", "Téma"])
    r += 1

    schedule_data = [
        ("Pondělí",  "Reel",     "Vzdělávání",          "Workout tip nebo fitness fakt",              COLOR_EDU),
        ("Úterý",    "Carousel", "Vzdělávání",          "5 tipů, checklist, mýtus vs. pravda",        COLOR_EDU),
        ("Středa",   "Reel",     "Humor",               "Relatable gymová situace, meme",             COLOR_HUM),
        ("Čtvrtek",  "Carousel", "Appka / Social proof","App feature nebo member story",              COLOR_APP),
        ("Pátek",    "Reel",     "Vzdělávání / Humor",  "Volba týmu — co teď letí (/suggest pondělí)",COLOR_EDU),
        ("Sobota",   "Reel",     "Collab / B2B",        "Gymový collab nebo obsah pro majitele posiloven", COLOR_APP),
        ("Neděle",   "—",        "Stories only",        "Zákulisí, poll, Q&A, repost gymů",          COLOR_OFF),
    ]
    for den, fmt, typ, tema, color in schedule_data:
        set_row_height(ws, r, 32)
        data_cell(ws, r, 1, den, bg=color, bold=True)
        data_cell(ws, r, 2, fmt, bg=color, align="center", bold=(fmt=="Reel"))
        data_cell(ws, r, 3, typ, bg=color)
        data_cell(ws, r, 4, tema, bg=color)
        r += 1

    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1,
        value="⚠️  Květen 2026: pouze 3 Reels/týden (Po + St + Pá). Carousel Út a Čt + Reel So přibývají od 1. 6. 2026.")
    c.fill = fill(YELLOW)
    c.font = Font(bold=True, color=NAVY, size=10)
    c.alignment = left()
    set_row_height(ws, r, 22)
    r += 2

    # ── PRAVIDLA AGENTA ──
    section_title(ws, r, 1, "  MARKETING AGENT — PRAVIDLA A PŘÍKAZY", 4)
    r += 1
    header_row(ws, r, ["Situace", "Příkaz", "Kdo spouští", "Frekvence"])
    r += 1

    agent_data = [
        ("Každý nový post",           "/post + brief nebo foto",  "Petr / Richard", "Každý den"),
        ("Série 5+ postů (kampaň)",   "/campaign + brief",        "Petr + David",   "Měsíčně"),
        ("Zjistit co teď letí",       "/suggest",                 "Kdokoli",        "Každé pondělí"),
        ("Výsledky za týden",         "/report",                  "Automaticky",    "Každý pátek"),
        ("Blog / SEO článek",         "/blog + téma",             "Richard",        "2× měsíčně"),
    ]
    for sit, prik, kdo, freq in agent_data:
        set_row_height(ws, r, 28)
        data_cell(ws, r, 1, sit)
        data_cell(ws, r, 2, prik, bg=GRAY_LIGHT, bold=True)
        data_cell(ws, r, 3, kdo, align="center")
        data_cell(ws, r, 4, freq, align="center")
        r += 1
    r += 1

    # ── BRAND GUIDELINES ──
    section_title(ws, r, 1, "  BRAND GUIDELINES", 4)
    r += 1
    brand_data = [
        ("Primární barva",    "#0B1222 (Dark Navy)",    "Pozadí, texty, nadpisy"),
        ("Sekundární barva",  "#4CC9FF (Electric Blue)", "Akcenty, CTA, highlights"),
        ("Font",              "Inter / Montserrat",      "Všechny vizuály a Reels"),
        ("Tón",               "Přímý, motivující, autentický", "Žádný marketing bullshit"),
        ("Čísla",             "NIKDY nevymýšlet",        "Jen: '30–50 % odchody' (ověřeno z 32 rozhovorů)"),
        ("Ceny",              "Do obsahu NEPATŘÍ",       "Ani B2C ani B2B ceny v postech"),
        ("Jazyk",             "Výhradně čeština",        "Do konce roku 2026"),
        ("Schvalování",       "Vždy David schvaluje",    "Žádný post jde live bez 'ok' od Davida"),
    ]
    header_row(ws, r, ["Prvek", "Hodnota", "Poznámka"], bg=GRAY_DARK)
    r += 1
    for prvek, hodnota, pozn in brand_data:
        set_row_height(ws, r, 26)
        data_cell(ws, r, 1, prvek, bold=True)
        data_cell(ws, r, 2, hodnota, bg=GRAY_LIGHT)
        data_cell(ws, r, 3, pozn)
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — KALENDÁŘ
# ══════════════════════════════════════════════════════════════════════════════
def make_kalendar(wb):
    ws = wb.create_sheet("📅 Kalendář")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    col_widths = [10, 11, 22, 12, 18, 30, 16, 26, 14, 50, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:K1")
    c = ws.cell(row=1, column=1, value="📅  PUMPLO — CONTENT KALENDÁŘ  |  Máj–Prosinec 2026")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.alignment = center()
    set_row_height(ws, 1, 36)

    headers = [
        "Datum", "Den", "Fáze", "Formát", "Typ obsahu (%)",
        "Téma / Brief", "Kdo natáčí", "Hashtagy (bank)",
        "Status", "Caption (vygeneroval agent)", "Poznámky"
    ]
    header_row(ws, 2, headers, height=32)

    STATUS_OPTIONS = ["📝 Naplánováno", "🎬 Natáčí se", "✏️ Agent píše caption",
                      "👀 Čeká na schválení", "✅ Zveřejněno", "⏭️ Přeskočeno"]

    r = 3
    start = date(2026, 5, 11)   # první post v pondělí 11. 5.
    end   = date(2026, 12, 31)
    d = start
    while d <= end:
        weekday = d.weekday()
        phase = get_phase(d)
        phase_color = PHASE_COLORS[phase].replace("#", "")
        poznamka = ""

        # ── Special days override ─────────────────────────────────────────────
        if d in SPECIAL_DAYS:
            fmt, typ, tema_default, content_color, poznamka = SPECIAL_DAYS[d]

        else:
            fmt_info = WEEKLY[weekday]
            fmt, typ, tema_default, content_color = fmt_info

            # Fáze 0 (normální dny): jen 3 Reely/týden (Po + St + Pá), zbytek Stories
            if phase == "Fáze 0 — Rozjezd" and weekday in PHASE0_SKIP:
                fmt = "—"
                typ = "Stories only"
                tema_default = "Zákulisí, poll, Q&A"
                content_color = COLOR_OFF

            if fmt is None:
                fmt = "—"

        is_launch = d == date(2026, 6, 1)
        is_first  = d == date(2026, 5, 11)
        row_h = 40 if (is_launch or is_first) else 30

        set_row_height(ws, r, row_h)

        data_cell(ws, r, 1,  d.strftime("%d. %m. %Y"),
                  bg=content_color, align="center",
                  bold=(is_launch or is_first))
        data_cell(ws, r, 2,  CZ_DAYS[weekday],            bg=content_color)
        data_cell(ws, r, 3,  phase,                        bg=phase_color)
        data_cell(ws, r, 4,  fmt,
                  bg=content_color, bold=(fmt in ("Reel", "🚀 LAUNCH")), align="center")
        data_cell(ws, r, 5,  typ,                          bg=content_color)
        data_cell(ws, r, 6,  tema_default,                 bg=WHITE)
        data_cell(ws, r, 7,  "Simona / Samuel / Trenér",   bg=GRAY_LIGHT)
        data_cell(ws, r, 8,  "viz záložka Hashtagy",       bg=GRAY_LIGHT)
        data_cell(ws, r, 9,  "📝 Naplánováno",             bg=WHITE)
        data_cell(ws, r, 10, "",                            bg=WHITE)
        data_cell(ws, r, 11, poznamka,                     bg=(YELLOW if poznamka else WHITE))

        r += 1
        d += timedelta(days=1)

    # Auto-filter
    ws.auto_filter.ref = f"A2:K{r-1}"


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — BACKLOG NÁPADŮ
# ══════════════════════════════════════════════════════════════════════════════
def make_backlog(wb):
    ws = wb.create_sheet("✅ Backlog nápadů")
    ws.sheet_view.showGridLines = False

    col_widths = [40, 12, 20, 12, 16, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="✅  BACKLOG NÁPADŮ — sem přidává kdokoli z týmu nebo bot (/suggest)")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=13)
    c.alignment = center()
    set_row_height(ws, 1, 34)

    headers = ["Nápad / Téma", "Formát", "Typ obsahu", "Priorita", "Navrhl", "Datum", "Poznámky"]
    header_row(ws, 2, headers)

    seed_ideas = [
        ("3 důvody proč 50 % lidí opustí posilovnu po 6 měsících",     "Reel",     "Vzdělávání",        "🔴 Vysoká",   "David",   "6. 5. 2026",  "Statistika z validace — 243 členů"),
        ("POV: Přijdeš do gymu a trenér ví co chceš cvičit",           "Reel",     "Humor",             "🔴 Vysoká",   "Petr",    "6. 5. 2026",  "Relatable, virální potenciál"),
        ("5 věcí které kvalitní posilovna dělá jinak",                  "Carousel", "Vzdělávání",        "🔴 Vysoká",   "Richard", "6. 5. 2026",  "B2B obsah pro majitele"),
        ("Behind the scenes: jak Pumplo vzniklo",                       "Reel",     "Behind the scenes", "🟡 Střední",  "David",   "6. 5. 2026",  "Founder story"),
        ("Tvůj tréninkový plán — jak to funguje v Pumplo",             "Carousel", "Appka",             "🔴 Vysoká",   "Simona",  "6. 5. 2026",  "App showcase — screenshoty"),
        ("Trenér z Eurogymu: 'Tohle změnilo mou práci'",               "Reel",     "Social proof",      "🟡 Střední",  "Petr",    "6. 5. 2026",  "Collab s trenérem z Eurogymu"),
        ("Mýtus vs. Pravda: Čím více tréninků tím lépe?",              "Carousel", "Vzdělávání",        "🟡 Střední",  "Richard", "6. 5. 2026",  "High saves potenciál"),
        ("Člen Eurogymu: 'Poprvé vím co mám cvičit'",                  "Reel",     "Social proof",      "🟡 Střední",  "Simona",  "6. 5. 2026",  "Testimonial — natočit v gymu"),
        ("Jak vypadá den s Pumplo — morning routine",                   "Reel",     "Appka",             "🟢 Nízká",    "David",   "6. 5. 2026",  "Lifestyle obsah"),
        ("Data z posilovny: Co dělají úspěšné gymy jinak",             "Carousel", "Vzdělávání/B2B",    "🟢 Nízká",    "Richard", "6. 5. 2026",  "Myzone/Mindbody benchmark data"),
        ("6 cviků které děláš špatně (a jak je opravit)",              "Carousel", "Vzdělávání",        "🟡 Střední",  "Trenér",  "6. 5. 2026",  "Natočit v Europymu"),
        ("Launch day: Pumplo je live v 5 posilovnách! 🎉",             "Reel",     "Behind the scenes", "🔴 Vysoká",   "David",   "1. 6. 2026",  "Launch post — připravit v předstihu"),
    ]

    r = 3
    priority_colors = {
        "🔴 Vysoká": "FFD0D0",
        "🟡 Střední": "FFF5CC",
        "🟢 Nízká": "D0FFD0",
    }
    for idea_data in seed_ideas:
        napad, fmt, typ, priorita, navrhl, datum, pozn = idea_data
        color = priority_colors.get(priorita, WHITE)
        set_row_height(ws, r, 30)
        data_cell(ws, r, 1, napad)
        data_cell(ws, r, 2, fmt, align="center")
        data_cell(ws, r, 3, typ)
        data_cell(ws, r, 4, priorita, bg=color, bold=True, align="center")
        data_cell(ws, r, 5, navrhl, align="center")
        data_cell(ws, r, 6, datum, align="center")
        data_cell(ws, r, 7, pozn, bg=GRAY_LIGHT)
        r += 1

    ws.auto_filter.ref = f"A2:G{r-1}"


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — HASHTAGY
# ══════════════════════════════════════════════════════════════════════════════
def make_hashtagy(wb):
    ws = wb.create_sheet("#️⃣ Hashtagy")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30

    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1,
        value="#️⃣  HASHTAGY — vždy 2–3 velké + 3–4 střední + 2–3 niche = 8–10 celkem")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=13)
    c.alignment = center()
    set_row_height(ws, 1, 34)

    header_row(ws, 2, ["Skupina", "Hashtagy", "Poznámka"])

    hashtag_data = [
        ("🔴 VELKÉ (>500k)", "#fitness #posilovna #trening #gym #workout #fitnessmotivace",
         "Velký dosah, ale velká konkurence — max 3 na post", "FFD0D0"),
        ("🟡 STŘEDNÍ (50k–500k)", "#ceskyfit #workoutcz #gymlife #fitnesscz #cviceni #zdravylivotni #pohyb #fitnessinspiration",
         "Zlatá střední cesta — hlavní záběr", "FFF5CC"),
        ("🟢 NICHE (pod 50k)", "#pumplo #nezavislagyma #gymretence #fitnesssoftware #posilovnacz #personaltraining #trenerfit #sportcz",
         "Nízká konkurence — Pumplo se zobrazí na prvním místě", "D0FFD0"),
        ("💙 B2B specifické", "#gymowner #fitnessbusiness #gymmanagement #boutiquefitness #fitnessapp #retenceclenu",
         "Použít v sobotním B2B postu nebo Carousel pro majitele", COLOR_APP),
        ("😂 Humor posty", "#gymhumor #gymproblems #fitnessmemes #posilovna #gymlife #totoznate",
         "Pro středeční humor Reel", COLOR_HUM),
        ("🏋️ Vzdělávání", "#workoutoftheday #fitnessfakta #treninkovyplan #zdravytrenink #personaltraining",
         "Pro pondělní a páteční vzdělávací Reels", COLOR_EDU),
        ("🚀 Launch hashtagy", "#pumplo #PumploFitness #MůjPumplo #pumplo_official",
         "Branded hashtag — push od launch dne 1. 6. 2026", COLOR_BTS),
    ]

    r = 3
    for skupina, hts, pozn, color in hashtag_data:
        set_row_height(ws, r, 50)
        data_cell(ws, r, 1, skupina, bg=color, bold=True)
        data_cell(ws, r, 2, hts, bg=color)
        data_cell(ws, r, 3, pozn, bg=GRAY_LIGHT)
        r += 1

    r += 1
    section_title(ws, r, 1, "  DOPORUČENÉ KOMBINACE PRO RYCHLÝ VÝBĚR", 3)
    r += 1
    header_row(ws, r, ["Typ postu", "Doporučená kombinace (kopíruj a uprav)", ""])
    r += 1

    combos = [
        ("Reel — Vzdělávání (Po/Pá)",
         "#fitness #ceskyfit #workoutcz #treninkovyplan #fitnessfakta #posilovna #pumplo #zdravytrenink"),
        ("Reel — Humor (St)",
         "#gymhumor #gymlife #posilovna #fitness #totoznate #fitnessmemes #ceskyfit #pumplo"),
        ("Carousel — Vzdělávání (Út)",
         "#fitness #ceskyfit #fitnessmotivace #zdravylivotni #workoutcz #treninkovyplan #pumplo #nezavislagyma"),
        ("Reel — Collab / B2B (So)",
         "#gymowner #fitnessbusiness #posilovna #gymmanagement #ceskyfit #fitness #pumplo #boutiquefitness"),
        ("Appka showcase (Čt)",
         "#fitnessapp #fitnesscz #treninkovyplan #pumplo #PumploFitness #fitness #posilovna #gymlife"),
    ]
    for typ, combo in combos:
        set_row_height(ws, r, 40)
        data_cell(ws, r, 1, typ, bold=True)
        data_cell(ws, r, 2, combo, bg=GRAY_LIGHT)
        data_cell(ws, r, 3, "")
        r += 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — KPIs
# ══════════════════════════════════════════════════════════════════════════════
def make_kpis(wb):
    ws = wb.create_sheet("📈 KPIs")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value="📈  KPIs — Měsíční tracking: Cíl vs. Skutečnost")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.alignment = center()
    set_row_height(ws, 1, 36)

    months = ["Květen", "Červen", "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"]
    col_widths = [28] + [16] * (len(months) * 2)
    for i, w in enumerate(col_widths, 1):
        if i <= len(col_widths):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Month headers (merged pairs)
    r = 2
    ws.cell(row=r, column=1, value="Metrika").fill = fill(NAVY)
    ws.cell(row=r, column=1).font = Font(bold=True, color=WHITE, size=11)
    ws.cell(row=r, column=1).alignment = center()
    ws.cell(row=r, column=1).border = thin_border()
    set_row_height(ws, r, 28)

    for i, month in enumerate(months):
        col_start = 2 + i * 2
        ws.merge_cells(start_row=r, start_column=col_start,
                       end_row=r, end_column=col_start + 1)
        c = ws.cell(row=r, column=col_start, value=month)
        c.fill = fill(BLUE)
        c.font = Font(bold=True, color=NAVY, size=11)
        c.alignment = center()
        c.border = thin_border()

    r += 1
    set_row_height(ws, r, 22)
    ws.cell(row=r, column=1, value="").border = thin_border()
    for i in range(len(months)):
        col_start = 2 + i * 2
        for col, label in [(col_start, "🎯 Cíl"), (col_start + 1, "✅ Skutečnost")]:
            c = ws.cell(row=r, column=col, value=label)
            c.fill = fill(GRAY_LIGHT)
            c.font = Font(bold=True, size=9)
            c.alignment = center()
            c.border = thin_border()

    r += 1

    # Target values per month
    targets = {
        "Followers (celkem)":       [200,   600,  1000, 1500, 2500, 4000, 7000, 10000],
        "Nové followers za měsíc":  [185,   400,   400,  500, 1000, 1500, 3000,  3000],
        "Reels plays (průměr)":     [500,  1500,  2000, 3000, 5000, 8000,12000, 20000],
        "Engagement rate %":        ["2%",  "3%",  "3%", "4%", "4%", "4%", "5%",  "5%"],
        "Reach (měsíční)":          [2000, 8000, 12000,18000,30000,50000,80000,120000],
        "Stories views (průměr)":   [50,    200,   300,  400,  600, 1000, 2000,  4000],
        "Počet Reels":              [12,    20,    20,   22,   22,   22,   22,    22],
        "Počet Carousels/Fotek":    [5,     10,    10,   10,   10,   10,   10,    10],
        "Collab posty":             [1,      4,     4,    4,    4,    4,    4,     4],
    }

    bg_alt = [WHITE, GRAY_LIGHT]
    for idx, (metrika, values) in enumerate(targets.items()):
        bg = bg_alt[idx % 2]
        set_row_height(ws, r, 26)
        data_cell(ws, r, 1, metrika, bg=bg, bold=True)
        for i, val in enumerate(values):
            col_start = 2 + i * 2
            data_cell(ws, r, col_start, val, bg=bg, align="center", bold=True)
            data_cell(ws, r, col_start + 1, "", bg=WHITE, align="center")  # skutečnost
        r += 1

    r += 1
    # Legend
    ws.merge_cells(f"A{r}:J{r}")
    c = ws.cell(row=r, column=1,
        value="ℹ️  Aktualizuj každý pátek po /report od agenta. Engagement rate benchmark pro fitness: 2–4 %. Followers pod cílem = přidat Collab post nebo /campaign.")
    c.fill = fill(YELLOW)
    c.font = Font(bold=False, color=NAVY, size=10)
    c.alignment = left()
    set_row_height(ws, r, 28)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — BOT GUIDE
# ══════════════════════════════════════════════════════════════════════════════
def make_bot_guide(wb):
    ws = wb.create_sheet("🤖 Bot guide")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 18

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="🤖  MARKETING AGENT — Jak ho použít (Telegram)")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.alignment = center()
    set_row_height(ws, 1, 36)

    ws.merge_cells("A2:D2")
    c = ws.cell(row=2, column=1,
        value="Agent generuje caption, hashtagy a CTA za ~30 sekund. David vždy schvaluje před publikováním. Nikdy nevymýšlí čísla.")
    c.fill = fill(BLUE)
    c.font = Font(bold=True, color=NAVY, size=11)
    c.alignment = center()
    set_row_height(ws, 2, 24)

    r = 3

    # Workflow
    section_title(ws, r, 1, "  WORKFLOW — každý post", 4)
    r += 1
    steps = [
        ("1️⃣", "Simona natočí / Petr nebo Richard mají nápad",
         "Brief = 1–3 věty o čem je post. Nebo přilož fotku/video.", ""),
        ("2️⃣", "Napiš agentovi na Telegram: /post [brief]",
         "Příklad: '/post Reel o tom proč lidé odcházejí z gymu po 3 měsících, humor, 20s'", ""),
        ("3️⃣", "Agent vrátí: caption, hashtagy, CTA návrh",
         "Za ~30 sekund. Upravit můžeš kdykoliv.", ""),
        ("4️⃣", "Simona přidá titulky přes Captions app",
         "Automatické titulky — povinné pro každý Reel", ""),
        ("5️⃣", "Pošli náhled Davidovi k schválení",
         "David napíše 'ok' → Ayrshare publikuje na IG (+ FB)", ""),
    ]
    header_row(ws, r, ["Krok", "Akce", "Detail", "Kdo"])
    r += 1
    for krok, akce, detail, kdo in steps:
        set_row_height(ws, r, 40)
        data_cell(ws, r, 1, krok, align="center", bold=True, bg=GRAY_LIGHT)
        data_cell(ws, r, 2, akce, bold=True)
        data_cell(ws, r, 3, detail, bg=GRAY_LIGHT)
        data_cell(ws, r, 4, kdo, align="center")
        r += 1

    r += 1
    section_title(ws, r, 1, "  PŘÍKAZY AGENTA — referenční tabulka", 4)
    r += 1
    header_row(ws, r, ["Příkaz", "Kdy použít", "Příklad briefe", "Frekvence"])
    r += 1

    commands = [
        ("/post",     "Každý nový příspěvek",
         "'/post Carousel 5 důvodů proč posilovat ráno, vzdělávání, B2C'",
         "Každý den"),
        ("/campaign", "Plánovat sérii 5+ postů najednou",
         "'/campaign Červnový launch — 6 postů o spuštění appky v gymech'",
         "Měsíčně"),
        ("/suggest",  "Zjistit co teď letí ve fitness na IG",
         "'/suggest — co jsou teď trendy témata v české fitness komunitě?'",
         "Každé pondělí"),
        ("/report",   "Výsledky za uplynulý týden",
         "'/report' — agent vrátí dosah, engagement, top post, doporučení",
         "Každý pátek"),
        ("/blog",     "SEO článek na web Pumplo",
         "'/blog Jak snížit churn v posilovně o 30 % — 5 strategií'",
         "2× měsíčně"),
    ]
    for cmd, kdy, priklad, freq in commands:
        set_row_height(ws, r, 44)
        data_cell(ws, r, 1, cmd, bold=True, bg=GRAY_LIGHT, align="center")
        data_cell(ws, r, 2, kdy)
        data_cell(ws, r, 3, priklad, bg=GRAY_LIGHT)
        data_cell(ws, r, 4, freq, align="center")
        r += 1

    r += 1
    section_title(ws, r, 1, "  TIPY PRO LEPŠÍ VÝSLEDKY", 4)
    r += 1

    tips = [
        ("Brief je zlatý", "Čím konkrétnější brief, tím lepší caption. Napiš: formát, téma, cílová skupina, tón (humor/vzdělávání/inspirace)."),
        ("Hashtagy", "Vždy použij kombinaci z záložky 'Hashtagy'. Bot hashtagy navrhne, ale zkontroluj že jsou aktuální."),
        ("Titulky (Captions)", "VŽDY přidej titulky na Reels. 85 % lidí sleduje videa bez zvuku. Captions app to udělá automaticky."),
        ("Timing", "Nejlepší čas publikovat: Út–Pá 7:00–9:00 nebo 17:00–20:00. Bot může nastavit scheduled post přes Ayrshare."),
        ("Collab posty", "Pokud natáčíš v partnerském gymu, použij IG Collab feature — post se zobrazí i na jejich profilu."),
        ("Virální hook", "Prvních 3 sekundy rozhodují. Brief vždy zahrň: 'hook má být: [konkrétní věta]' — agent ho zakomponuje do caption a titulků."),
        ("Čísla", "NIKDY nevymýšlej čísla. Používej jen: '30–50 % členů odchází' (ověřeno z 32 rozhovorů + 243 member surveys)."),
        ("Schválení", "Davidovi pošli: 1) caption, 2) hashtagy, 3) náhled videa nebo foto. On odpoví 'ok' nebo s úpravami."),
    ]
    header_row(ws, r, ["Tip", "Detail", "", ""])
    r += 1
    for tip, detail in tips:
        set_row_height(ws, r, 36)
        data_cell(ws, r, 1, tip, bold=True, bg=GRAY_LIGHT)
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r, end_column=4)
        data_cell(ws, r, 2, detail)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(row=r, column=1,
        value="📞 Kontakt: info.pumplo@gmail.com  |  IG: @pumplo_official  |  Web: pumplo.com")
    c.fill = fill(NAVY)
    c.font = Font(bold=True, color=BLUE, size=10)
    c.alignment = center()
    set_row_height(ws, r, 24)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    wb = build_workbook()
    out_path = os.path.join(
        os.path.dirname(__file__), "..",
        "Pumplo_Instagram_Strategie_2026.xlsx"
    )
    out_path = os.path.abspath(out_path)
    wb.save(out_path)
    print(f"✅  Soubor uložen: {out_path}")
